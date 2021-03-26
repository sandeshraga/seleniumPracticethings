import openpyxl
#path for excel file
path="D:\Study material\MS\selenium testing codes\excel operations\practice book.xlsx"

workbook=openpyxl.load_workbook(path)
#to select active sheet from excel file
sheet=workbook.active #workbook.get_sheet_by_name("sheet1")

rows=sheet.max_row #returns no. of rows
cols=sheet.max_column #returns no. of column

print(rows)
print(cols)

#to print each cell data
for r in range(1,rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r,column=c).value, end="      ")
    print()