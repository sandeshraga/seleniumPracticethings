import openpyxl

def get_Row_Count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)

def get_column_Count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)

def read_data_file(file,sheetname,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value

def write_data_file(file,sheetname,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value=data
    workbook.save(file)
